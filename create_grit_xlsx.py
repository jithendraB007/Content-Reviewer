"""
Creates grit_l1_input.xlsx directly from the GRIT L1 Set 4 CSV data.
Run: python create_grit_xlsx.py
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

RAW_DATA = [
    # (Instruction, Question, Options list, Correct Answer, Explanation, Skills Tested, Difficulty, Sub-type)
    ("Choose the word that best matches the meaning of the highlighted word.", "The meeting was quick.", ["fast","slow","long","late"], "fast", "'Quick' means something done in a short time.", "Vocabulary in context", "A1", "Synonym"),
    ("Choose the correct opposite meaning of the word.", "Select the antonym of happy.", ["sad","glad","smile","joy"], "sad", "'Sad' is the opposite of 'happy.'", "Synonyms & Antonyms", "A1", "Antonym"),
    ("Choose the correct word pair to complete the sentence.", "She made a cup of ____ tea.", ["hot","warm","cold","hard"], "hot", "'Hot tea' is a common word pair.", "Common Word Pairs", "A1", "Collocation"),
    ("Read the sentence and identify if it is a fact or opinion.", "'Dogs are better than cats.'", ["Fact","Opinion","Instruction","Question"], "Opinion", "This is a personal belief, not a proven fact.", "Fact vs Opinion", "A1", "Classification"),
    ("Choose the correct phrase to complete the sentence.", "_____ please pass the file?", ["Can you","You can","Can me","Me can"], "Can you", "'Can you' is a polite request.", "Functional Phrases", "A1", "Request"),
    ("Choose the correct answer based on logic.", "Ravi is taller than Aman. Aman is taller than Raj. Who is tallest?", ["Ravi","Aman","Raj","None"], "Ravi", "Ravi is taller than both.", "Simple Reasoning", "A1", "Conclusion"),
    ("Choose the correct meaning of the word.", "The room is clean.", ["dirty","neat","messy","bad"], "neat", "'Clean' means neat and tidy.", "Vocabulary in context", "A1", "Synonym"),
    ("Choose the correct opposite meaning.", "Select the antonym of early.", ["soon","late","fast","quick"], "late", "'Late' is opposite of 'early.'", "Synonyms & Antonyms", "A1", "Antonym"),
    ("Choose the correct phrase.", "He is ____ his homework.", ["doing","making","taking","giving"], "doing", "'Doing homework' is the correct phrase.", "Common Word Pairs", "A1", "Collocation"),
    ("Choose the correct conclusion.", "It is raining. What should you carry?", ["book","umbrella","pen","bag"], "umbrella", "Rain requires an umbrella.", "Simple Reasoning", "A1", "Inference"),
    ("Choose the word that best matches the meaning of the highlighted word.", "The manager gave a clear instruction to the team.", ["easy","short","simple","plain"], "plain", "'Clear' here means easy to understand or plain in meaning.", "Vocabulary in context", "A2", "Synonym"),
    ("Choose the correct opposite meaning of the word.", "Select the antonym of borrow.", ["lend","take","keep","hold"], "lend", "To 'borrow' means to take something for a short time, while to 'lend' means to give it.", "Synonyms & Antonyms", "A2", "Antonym"),
    ("Choose the correct word pair to complete the sentence.", "Please _____ a seat while you wait.", ["do","take","bring","hold"], "take", "'Take a seat' is the correct and common expression.", "Common Word Pairs", "A2", "Collocation"),
    ("Read the statement and identify it correctly.", "'The Earth moves around the Sun.'", ["Fact","Opinion","Advice","Request"], "Fact", "This is a scientifically proven statement, so it is a fact.", "Fact vs Opinion", "A2", "Classification"),
    ("Choose the most suitable phrase for the situation.", "You want your classmate to lower the volume. What will you say?", ["Turn it off now.","Can you please lower the volume?","You are wrong.","I do not like music."], "Can you please lower the volume?", "This is the most polite and appropriate request.", "Functional Phrases", "A2", "Polite Request"),
    ("Choose the most logical conclusion.", "Sonia left home with an umbrella and a raincoat. What can you infer?", ["She is going shopping.","She expects wet weather.","She is going to school.","She forgot her bag."], "She expects wet weather.", "An umbrella and a raincoat suggest that she expects rain or wet weather.", "Simple Conclusion", "A2", "Inference"),
    ("Choose the word that best matches the meaning of the highlighted word.", "The shop is closed today.", ["open","shut","busy","near"], "shut", "'Closed' means not open for use or service; 'shut' is the closest meaning.", "Vocabulary in context", "A2", "Synonym"),
    ("Choose the correct opposite meaning of the word.", "Select the antonym of empty.", ["vacant","full","clean","open"], "full", "'Full' is the opposite of 'empty.'", "Synonyms & Antonyms", "A2", "Antonym"),
    ("Choose the correct phrase to complete the sentence.", "The teacher asked the students to _____ attention.", ["do","take","pay","keep"], "pay", "'Pay attention' is the correct expression.", "Common Word Pairs", "A2", "Collocation"),
    ("Choose the best answer based on the information given.", "Neha reached the station at 8:00. The train left at 7:45. What is the most likely result?", ["She caught the train.","She missed the train.","The train was delayed.","She bought a ticket."], "She missed the train.", "The train left before Neha arrived, so she most likely missed it.", "Simple Conclusion", "A2", "Logical Outcome"),
    # Reading Comprehension - A1
    ("Read the passage and choose the correct answer.", "Rina goes to school by bus every day. She leaves home at 7:30 in the morning and reaches school at 8:00. She likes travelling by bus because it is easy and cheap.\n\nHow does Rina go to school?", ["By car","By bus","By train","On foot"], "By bus", "The passage clearly states that Rina goes to school by bus every day.", "Main idea and specific detail", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Ali has a small garden behind his house. He waters the plants every evening and removes dry leaves on Sundays. His favourite flowers are roses.\n\nWhat does Ali do every evening?", ["He buys flowers.","He waters the plants.","He cleans the house.","He cuts the trees."], "He waters the plants.", "The passage says that Ali waters the plants every evening.", "Locating specific details", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Meena was late for class because her bicycle had a flat tyre. She walked to school, but it took longer than usual.\n\nWhy was Meena late for class?", ["She woke up late.","Her bicycle had a flat tyre.","She missed the bus.","She forgot her books."], "Her bicycle had a flat tyre.", "The passage directly states the reason for her delay.", "Cause and effect", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Rohit studied hard for his maths test. He solved many practice questions and revised his notes. On the test day, he felt confident.\n\nHow did Rohit feel on the test day?", ["Nervous","Confident","Angry","Tired"], "Confident", "The passage says that he felt confident on the test day.", "Locating specific details", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Sara packed a bottle of water, a cap, and some fruit before leaving for the park. The day was sunny and warm.\n\nWhy did Sara pack these things?", ["She was going to the market.","She was preparing for warm weather.","She was going to a library.","She wanted to stay at home."], "She was preparing for warm weather.", "The items she packed fit a sunny and warm day outdoors.", "Simple inference", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "The school library opens at 9:00 a.m. and closes at 4:00 p.m. Students can borrow two books at a time for one week.\n\nHow many books can a student borrow at one time?", ["One","Two","Three","Four"], "Two", "The passage clearly says students can borrow two books at a time.", "Locating specific details", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Arun forgot to charge his phone at night. In the morning, the battery was low, so he could not listen to music on the way to work.\n\nWhat caused the problem in the morning?", ["He lost his phone.","He forgot to charge it.","He missed the bus.","He dropped the phone."], "He forgot to charge it.", "The low battery happened because he did not charge the phone the previous night.", "Cause and effect", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Leena enjoys cooking with her mother on weekends. They usually make rice, vegetables, and soup together. She says cooking makes her happy.\n\nWhat is the main idea of the passage?", ["Leena dislikes weekends.","Leena enjoys cooking with her mother.","Leena wants to buy a new pan.","Leena only likes soup."], "Leena enjoys cooking with her mother.", "This sentence captures the central idea of the short passage.", "Main idea identification", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "The office printer stopped working in the afternoon. The staff checked the paper tray and found that it was empty. After adding paper, the printer worked again.\n\nWhy did the printer stop working?", ["It had no paper.","It was switched off.","It was broken.","It had no ink."], "It had no paper.", "The staff found the paper tray empty, which caused the problem.", "Cause and effect", "A1", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Nita reads for thirty minutes before going to bed every night. She says this habit helps her relax and sleep better.\n\nWhat is one benefit of Nita's reading habit?", ["It helps her wake up early.","It helps her sleep better.","It helps her cook dinner.","It helps her finish homework."], "It helps her sleep better.", "The passage states that reading before bed helps her relax and sleep better.", "Key point extraction", "A1", "Reading Comprehension"),
    # Reading Comprehension - A2
    ("Read the passage and choose the correct answer.", "Mr. Das takes the metro to work because it is faster than driving in city traffic. He leaves home at 8:15 every morning and usually reaches his office before 9:00. He also likes the metro because he can read the news on the way.\n\nWhy does Mr. Das prefer the metro?", ["It is cheaper than all other transport.","It is faster than driving in traffic.","It stops near his friend's house.","It allows him to sleep on the way."], "It is faster than driving in traffic.", "The passage clearly states that he takes the metro because it is faster than driving in city traffic.", "Locating specific details", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Priya joined a spoken English class last month. She attends the class three times a week and practises at home every evening. Now she feels more confident when speaking to new people.\n\nWhat is the result of Priya's practice?", ["She speaks less at home.","She feels more confident speaking.","She has stopped attending class.","She avoids meeting new people."], "She feels more confident speaking.", "The passage says that Priya now feels more confident when speaking to new people.", "Cause and effect", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "The supermarket announced a discount on fruits and vegetables this weekend. Many customers visited early in the morning to buy fresh items before the shelves became empty.\n\nWhy did many customers visit early?", ["They wanted to meet the manager.","They wanted to buy fresh items first.","They had to return old products.","They were looking for household tools."], "They wanted to buy fresh items first.", "The passage explains that customers came early to buy fresh items before the shelves became empty.", "Simple inference", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "A school organised a cleanliness drive on Saturday. Students collected plastic bottles and paper from the playground and planted small trees near the gate. The principal thanked them for making the campus cleaner and greener.\n\nWhat did the students do during the drive?", ["They painted the classrooms.","They cleaned the playground and planted trees.","They repaired the school gate.","They visited another school."], "They cleaned the playground and planted trees.", "The passage directly states that the students collected waste and planted small trees.", "Main idea and specific detail", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Sameer wanted to submit his online form before the deadline. However, the internet connection at his home stopped working in the evening, so he went to a nearby computer centre and completed the form there.\n\nWhy did Sameer go to the computer centre?", ["He wanted to print a photo.","His home internet was not working.","He wanted to meet a friend.","His laptop battery was low."], "His home internet was not working.", "The passage says he went there because the internet connection at home stopped working.", "Cause and effect", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "The museum is open from Tuesday to Sunday and remains closed on Monday. Visitors must buy tickets at the front desk before entering. Photography is not allowed inside the main gallery.\n\nWhat must visitors do before entering the museum?", ["Take photographs.","Buy tickets at the front desk.","Visit only on Monday.","Show a school ID card."], "Buy tickets at the front desk.", "The passage clearly states that visitors must buy tickets before entering.", "Locating specific details", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Neelam keeps a small notebook where she writes down her daily expenses. At the end of each week, she checks the notebook to see how much money she has spent. This habit helps her plan her budget better.\n\nWhy does Neelam keep the notebook?", ["To practise handwriting every day.","To remember her friends' phone numbers.","To track spending and plan her budget.","To write stories at night."], "To track spending and plan her budget.", "The passage explains that she records expenses so that she can check spending and plan her budget.", "Key point extraction", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "During the science fair, Rohan presented a model of a water filter made with sand, stones, and charcoal. His teacher praised him because the model clearly showed how dirty water could be cleaned.\n\nWhy did the teacher praise Rohan?", ["He arrived early for the fair.","His model clearly explained water cleaning.","He helped arrange the chairs.","His model was the biggest one there."], "His model clearly explained water cleaning.", "The passage states that the teacher praised him because his model clearly showed how dirty water could be cleaned.", "Locating specific details", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "Maya set an alarm for 5:30 a.m. because she wanted to catch the first bus to her training centre. When the alarm rang, she got ready quickly and reached the bus stop on time.\n\nWhat can be understood from the passage?", ["Maya missed the first bus.","Maya reached the bus stop on time.","Maya forgot to set the alarm.","Maya walked to the training centre."], "Maya reached the bus stop on time.", "The passage says she got ready quickly and reached the bus stop on time.", "Simple inference", "A2", "Reading Comprehension"),
    ("Read the passage and choose the correct answer.", "A cafe introduced reusable cups for customers who sit inside the shop. This reduced the use of paper cups and helped the cafe produce less waste. Many regular customers appreciated this change.\n\nWhat was one result of introducing reusable cups?", ["The cafe became more expensive.","The use of paper cups decreased.","Customers stopped visiting the cafe.","The shop needed more chairs."], "The use of paper cups decreased.", "The passage clearly says that reusable cups reduced the use of paper cups.", "Cause and effect", "A2", "Reading Comprehension"),
    # Grammar - A1
    ("Select the correct verb form to complete the sentence.", "She ______ to school every day.", ["go","goes","going","gone"], "goes", "The subject 'She' is singular, so the verb takes 'goes' in the simple present tense.", "Subject-verb agreement", "A1", "Grammar"),
    ("Select the correct verb form to complete the sentence.", "They ______ happy after the test.", ["is","was","are","be"], "are", "The subject 'They' is plural, so 'are' is the correct form.", "Subject-verb agreement", "A1", "Grammar"),
    ("Select the correct preposition to complete the sentence.", "The class starts ______ 9 o'clock.", ["on","in","at","to"], "at", "'At' is used for clock times.", "Preposition usage", "A1", "Grammar"),
    ("Select the correct word form to complete the sentence.", "She sings very ______.", ["happy","happily","happiness","happier"], "happily", "An adverb is needed to describe how she sings, so 'happily' is correct.", "Word form usage", "A1", "Grammar"),
    ("Select the correct sentence.", "Choose the grammatically correct sentence.", ["He have a car.","He has a car.","He having a car.","He had a car."], "He has a car.", "'He has a car' is the correct present tense sentence.", "Simple sentence correction", "A1", "Grammar"),
    ("Select the correct modal to complete the sentence.", "You ______ drink water when you are thirsty.", ["fast","should","weak","late"], "should", "'Should' is used to give advice.", "Basic modal verbs", "A1", "Grammar"),
    ("Select the correct form to complete the sentence.", "I want ______ a new notebook.", ["take","taking","to take","taken"], "to take", "After 'want,' the infinitive form 'to take' is correct.", "Verb pattern usage", "A1", "Grammar"),
    ("Select the correct preposition to complete the sentence.", "The keys are ______ the table.", ["in","at","on","to"], "on", "'On' is used when something is placed on a surface.", "Preposition usage", "A1", "Grammar"),
    ("Arrange the sentences in the correct order to form a meaningful paragraph.", "1. She packed her books.\n2. Rina got ready for school.\n3. Then she left for the bus stop.\n4. First, she woke up early.", ["4, 2, 1, 3","2, 4, 1, 3","4, 1, 2, 3","1, 2, 4, 3"], "4, 2, 1, 3", "The logical order begins with waking up, then getting ready, packing books, and finally leaving.", "Sentence sequencing", "A1", "Grammar"),
    ("Select the best word to complete the short paragraph.", "Raj was late for class ______ he missed the bus.", ["and","but","so","because"], "because", "'Because' correctly shows the reason why Raj was late for class.", "Blank completion", "A1", "Grammar"),
    # Grammar - A2
    ("Select the correct verb form to complete the sentence.", "By the time the teacher entered the classroom, the students ______ their notebooks.", ["opened","open","had opened","have opened"], "had opened", "The action of opening the notebooks happened before another action in the past, so the past perfect form is correct.", "Present and past tense usage", "A2", "Grammar"),
    ("Select the correct verb form to complete the sentence.", "Last evening, the maintenance team ______ the lights in the hall before the event began.", ["checks","checked","checking","has checked"], "checked", "'Last evening' shows that the action happened in the past, so the simple past form is needed.", "Present and past tense usage", "A2", "Grammar"),
    ("Select the correct verb form to complete the sentence.", "Each of the students ______ given a separate worksheet for the activity.", ["have","were","was","are"], "was", "The subject 'Each' is singular, so it takes the singular verb 'was.'", "Subject-verb agreement", "A2", "Grammar"),
    ("Select the correct verb form to complete the sentence.", "The results of the practice test ______ available on the notice board now.", ["is","are","was","has been"], "are", "The subject 'results' is plural, so 'are' is the correct verb form.", "Subject-verb agreement", "A2", "Grammar"),
    ("Select the correct article to complete the sentence.", "She brought ______ umbrella because it looked like rain.", ["a","an","the","no article"], "an", "'Umbrella' begins with a vowel sound, so 'an' is the correct article.", "Article usage", "A2", "Grammar"),
    ("Select the correct article to complete the sentence.", "He is ______ honest employee who always completes his work on time.", ["a","an","the","no article"], "an", "'Honest' begins with a vowel sound because the 'h' is silent, so 'an' is correct.", "Article usage", "A2", "Grammar"),
    ("Select the correct preposition to complete the sentence.", "The workshop has been scheduled ______ Monday morning.", ["in","at","on","for"], "on", "'On' is used with days and specific parts of the day.", "Preposition usage", "A2", "Grammar"),
    ("Select the correct preposition to complete the sentence.", "The issue was discussed ______ the meeting and noted for follow-up.", ["on","in","about","for"], "in", "The natural expression is 'discussed in the meeting.'", "Preposition usage", "A2", "Grammar"),
    ("Select the correct modal to complete the sentence.", "The roads are wet, so you ______ drive carefully.", ["might","should","would","could have"], "should", "'Should' expresses advice or recommendation in this situation.", "Basic modal verbs", "A2", "Grammar"),
    ("Select the best correction for the sentence.", "Both of the boys is waiting outside the classroom.", ["Both of the boys is waiting outside the classroom.","Both of the boys are waiting outside the classroom.","Both of the boys were waiting outside the classroom.","Both of the boys has been waiting outside the classroom."], "Both of the boys are waiting outside the classroom.", "'Both' refers to two people and takes the plural verb 'are.'", "Simple sentence correction", "A2", "Grammar"),
    # Listening - A1
    ("Listen to the short conversation and choose the correct answer.", "Riya: 'Can you bring the blue file to the meeting room?' Aman: 'Sure, I will bring it in two minutes.'\n\nWhat does Aman say he will bring?", ["The red file","The blue file","The meeting notes","The laptop"], "The blue file", "Aman clearly says he will bring the blue file.", "Understanding short conversations", "A1", "Listening"),
    ("Listen to the instruction and choose the correct answer.", "Please open your notebook and write today's date at the top of the page.\n\nWhat should the listener do first?", ["Write a title","Open the notebook","Close the notebook","Read the page"], "Open the notebook", "The first instruction is to open the notebook.", "Following instructions", "A1", "Listening"),
    ("Listen to the sentence and choose the correct tone.", "Good morning. Please take your seat and wait for your turn.\n\nWhat is the speaker's tone?", ["Friendly","Angry","Excited","Funny"], "Friendly", "The speaker sounds polite and calm, which shows a friendly tone.", "Tone recognition", "A1", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Teacher: 'Where is your homework?' Student: 'I kept it in my bag, but I forgot to bring the bag today.'\n\nWhy does the student not have the homework?", ["The homework is unfinished","The student lost the homework","The student forgot the bag","The teacher took the homework"], "The student forgot the bag", "The student says the homework is in the bag, but the bag was forgotten.", "Understanding short conversations", "A1", "Listening"),
    ("Listen to the instruction and choose the correct answer.", "Turn off the fan and close the window before you leave the room.\n\nWhat should be done before leaving the room?", ["Open the fan and close the door","Turn off the fan and close the window","Close the window and switch on the light","Leave the room and close the window"], "Turn off the fan and close the window", "The instruction clearly gives these two actions before leaving.", "Following instructions", "A1", "Listening"),
    ("Listen to the sentence and choose the correct answer.", "The bus is late, so we may reach the office after 10 o'clock.\n\nWhat problem is mentioned?", ["The office is closed","The bus is late","The road is blocked","The meeting is cancelled"], "The bus is late", "The speaker directly says that the bus is late.", "Understanding short conversations", "A1", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Asha: 'Did you call the customer?' Ravi: 'Not yet. I will call after lunch.'\n\nWhen will Ravi call the customer?", ["In the morning","After lunch","Tomorrow","In the evening"], "After lunch", "Ravi clearly says he will call after lunch.", "Locating specific details", "A1", "Listening"),
    ("Listen to the sentence and choose the correct tone.", "Please do not worry. I will help you complete the form.\n\nWhat is the speaker's tone?", ["Rude","Supportive","Tired","Impatient"], "Supportive", "The speaker offers help in a reassuring way, so the tone is supportive.", "Tone recognition", "A1", "Listening"),
    ("Listen to the instruction and choose the correct answer.", "Take the ID card from the table and give it to the security guard.\n\nWho should receive the ID card?", ["The teacher","The manager","The security guard","The driver"], "The security guard", "The instruction says to give the ID card to the security guard.", "Following instructions", "A1", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Meena: 'Why are you carrying an umbrella?' Kiran: 'The sky is dark, and it may rain soon.'\n\nWhy is Kiran carrying an umbrella?", ["He is going to school","He thinks it may rain","He lost his bag","He wants to share it"], "He thinks it may rain", "Kiran says the sky is dark and it may rain soon.", "Simple inference", "A1", "Listening"),
    # Listening - A2
    ("Listen to the short conversation and choose the correct answer.", "Manager: 'Please send the updated report before lunch.' Employee: 'I have finished the first two sections, and I will complete the last section in the next thirty minutes.'\n\nWhat can be understood from the conversation?", ["The report has already been submitted.","The report is almost complete.","The report will be cancelled.","The report was finished yesterday."], "The report is almost complete.", "The employee says the first two sections are finished and only the last section is pending, so the report is almost complete.", "Understanding short conversations", "A2", "Listening"),
    ("Listen to the instruction and choose the correct answer.", "After you collect the forms from the front desk, place them on the manager's table and return to the training room.\n\nWhat should be done after collecting the forms?", ["Return to the training room immediately","Place them on the manager's table","Give them to the receptionist","Keep them in your bag"], "Place them on the manager's table", "The instruction clearly states that the forms must be placed on the manager's table after collection.", "Following instructions", "A2", "Listening"),
    ("Listen to the speaker and choose the correct tone.", "I understand that this delay is inconvenient, but we are doing everything possible to resolve the issue today.\n\nWhat is the speaker's tone?", ["Supportive","Angry","Uninterested","Confused"], "Supportive", "The speaker acknowledges the problem and offers reassurance, which shows a supportive tone.", "Tone recognition", "A2", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Student: 'I thought the test was tomorrow.' Teacher: 'It was moved to today because the school will remain closed tomorrow.'\n\nWhy was the test held today?", ["The teacher was absent yesterday","The school will be closed tomorrow","The students asked for an early test","The classroom was unavailable"], "The school will be closed tomorrow", "The teacher clearly states that the test was moved because the school will remain closed the next day.", "Locating specific details", "A2", "Listening"),
    ("Listen to the instruction and choose the correct answer.", "First switch off the projector, then close the laptop, and finally lock the conference room before you leave.\n\nWhat is the last action to be done?", ["Close the laptop","Leave the room","Lock the conference room","Switch off the projector"], "Lock the conference room", "The speaker says the final step is to lock the conference room.", "Following instructions", "A2", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Asha: 'Will you attend the online training at 4 PM?' Ravi: 'Yes, but I may join a few minutes late because my current meeting ends at the same time.'\n\nWhat can be concluded about Ravi?", ["He will miss the training completely","He may join the training a little late","He has already attended the training","He does not know about the training"], "He may join the training a little late", "Ravi says he may join a few minutes late because another meeting ends at the same time.", "Simple inference", "A2", "Listening"),
    ("Listen to the sentence and choose the correct answer.", "The package was delivered in the morning, but no one was available to receive it, so it was taken back.\n\nWhat happened to the package?", ["It was collected by the customer","It was left at the door","It was taken back","It was damaged in transit"], "It was taken back", "The speaker clearly says that the package was taken back because no one was available.", "Understanding spoken information", "A2", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Anita: 'Why are you leaving the office early today?' Karan: 'I have a doctor's appointment at 5 PM, so I need to leave by 4:30.'\n\nWhy is Karan leaving early?", ["He is not feeling well at work","He has to meet a client","He has a doctor's appointment","He has completed all his work"], "He has a doctor's appointment", "Karan directly states that he needs to leave early because of a doctor's appointment.", "Locating specific details", "A2", "Listening"),
    ("Listen to the speaker and choose the correct tone.", "Please remember to submit the travel form today. Without it, your reimbursement cannot be processed.\n\nWhat is the speaker's tone?", ["Serious","Humorous","Excited","Casual"], "Serious", "The speaker is giving an important reminder about a required form, which creates a serious tone.", "Tone recognition", "A2", "Listening"),
    ("Listen to the short conversation and choose the correct answer.", "Receptionist: 'Mr. Sharma is in a meeting at the moment. Could you please wait for ten minutes?' Visitor: 'Sure, I will wait in the lobby.'\n\nWhere will the visitor wait?", ["In the meeting room","At the front desk","In the lobby","Outside the building"], "In the lobby", "The visitor says he will wait in the lobby.", "Understanding short conversations", "A2", "Listening"),
]

def fmt_options(opts):
    labels = ["A", "B", "C", "D", "E"]
    return " | ".join(f"{labels[i]}) {o}" for i, o in enumerate(opts))

def get_qtype(sub_type, instruction):
    if sub_type == "Listening" or "listen" in instruction.lower():
        return "Audio Based MCQ"
    return "MCQ"

def get_transcript(sub_type, question):
    """For Listening questions, the script text before \n\n is the transcript."""
    if sub_type != "Listening":
        return ""
    parts = question.split("\n\n", 1)
    if len(parts) == 2:
        return parts[0].strip()
    return ""

def get_question_only(sub_type, question):
    if sub_type != "Listening":
        return question
    parts = question.split("\n\n", 1)
    return parts[1].strip() if len(parts) == 2 else question

rows = []
for i, (instr, question, opts, answer, expl, skills, diff, sub_type) in enumerate(RAW_DATA, 1):
    q_type = get_qtype(sub_type, instr)
    rows.append({
        "Q. NO": f"Q{i:03d}",
        "Question Type": q_type,
        "Transcript": get_transcript(sub_type, question),
        "Instructions": instr,
        "Question": get_question_only(sub_type, question),
        "Options": fmt_options(opts),
        "Correct Answer": answer,
        "Explanation": expl,
        "Schema": "",
        "Question Purpose": skills,
        "Difficulty": diff,
        "Tags": sub_type,
    })

df = pd.DataFrame(rows)
out_path = "grit_l1_input.xlsx"

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Questions")
    ws = writer.sheets["Questions"]

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    col_widths = {
        "A": 8, "B": 20, "C": 40, "D": 40, "E": 50,
        "F": 40, "G": 20, "H": 50, "I": 12, "J": 30, "K": 10, "L": 20,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Wrap text for all data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Created {out_path} with {len(df)} questions")
print(df["Question Type"].value_counts().to_string())
print(df["Tags"].value_counts().to_string())
