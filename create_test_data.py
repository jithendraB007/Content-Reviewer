"""
Run this script once to generate test_input.xlsx in the project root.
Usage: python create_test_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

COLUMNS = [
    "Q. NO", "Question Type", "Transcript", "Instructions", "Question",
    "Options", "Correct Answer", "Explanation", "Schema",
    "Question Purpose", "Difficulty", "Tags"
]

ROWS = [
    {
        "Q. NO": "Q001",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "Choose the correct answer.",
        "Question": "What is the capital of France?",
        "Options": "A) Paris | B) London | C) Berlin | D) Madrid",
        "Correct Answer": "A) Paris",
        "Explanation": "Paris is the capital and largest city of France.",
        "Schema": "Single correct",
        "Question Purpose": "Test knowledge of European capitals",
        "Difficulty": "Easy",
        "Tags": "Geography, Europe",
    },
    {
        "Q. NO": "Q002",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "Select the best option.",
        "Question": "Which of the following are a mammal?",
        "Options": "A) Eagle | B) Dolphin | C) Salmon | D) Cobra",
        "Correct Answer": "B) Dolphin",
        "Explanation": "Dolphins is a marine mammal unlike fish, birds and reptiles.",
        "Schema": "Single correct",
        "Question Purpose": "Classify animals by class",
        "Difficulty": "Medium",
        "Tags": "Animals, Biology",
    },
    {
        "Q. NO": "Q003",
        "Question Type": "Fill in the Blanks",
        "Transcript": "",
        "Instructions": "Fill in the blank.",
        "Question": "Water boils at ___ degrees at sea level.",
        "Options": "",
        "Correct Answer": "100",
        "Explanation": "Water undergoes phase transition from liquid to gas at 100 degrees Celsius under standard atmospheric pressure.",
        "Schema": "Single blank",
        "Question Purpose": "Temperature and phase changes",
        "Difficulty": "Easy",
        "Tags": "Physics, States of Matter",
    },
    {
        "Q. NO": "Q004",
        "Question Type": "Speaking Based",
        "Transcript": "",
        "Instructions": "Answer the following question aloud.",
        "Question": "Describe the main idea of the passage about climate change.",
        "Options": "",
        "Correct Answer": "The passage discusses rising temperatures and extreme weather",
        "Explanation": "A good answer should cover rising temperatures, sea level rise, and extreme weather events caused by climate change.",
        "Schema": "Open response",
        "Question Purpose": "Assess speaking and comprehension",
        "Difficulty": "Hard",
        "Tags": "Climate, Environment",
    },
    {
        "Q. NO": "Q005",
        "Question Type": "Audio Based MCQ",
        "Transcript": "The speaker talked about renewable energy sources including solar, wind, and hydro power as alternatives to fossil fuels.",
        "Instructions": "Listen to the audio and choose the correct answer.",
        "Question": "What is the most efficient form of renewable energy according to the speaker?",
        "Options": "A) Solar | B) Wind | C) Hydro | D) Geothermal",
        "Correct Answer": "A) Solar",
        "Explanation": "Renewable energy sources were mentioned but the speaker did not compare their efficiency.",
        "Schema": "Single correct",
        "Question Purpose": "Test listening comprehension",
        "Difficulty": "Medium",
        "Tags": "Energy, Environment",
    },
    {
        "Q. NO": "Q006",
        "Question Type": "Prompt Based",
        "Transcript": "",
        "Instructions": "Write a short paragraph.",
        "Question": "Write about the benefits of exercise. You kinda need to mention at least 3 points.",
        "Options": "",
        "Correct Answer": "Regular exercise improves physical and mental health.",
        "Explanation": "Exercise benefits include cardiovascular health, mental well-being, and weight management.",
        "Schema": "Open response",
        "Question Purpose": "Assess writing ability",
        "Difficulty": "Medium",
        "Tags": "Health, Fitness",
    },
    {
        "Q. NO": "Q007",
        "Question Type": "Image Based with Options",
        "Transcript": "",
        "Instructions": "Look at the image and choose the correct answer.",
        "Question": "What type of cloud formation is shown in the image",
        "Options": "A) Cumulus | B) Stratus | C) Cirrus | D) Nimbus",
        "Correct Answer": "A) Cumulus",
        "Explanation": "Cumulus clouds are characterised by their fluffy appearance. They organize into large formations.",
        "Schema": "Single correct",
        "Question Purpose": "Identify cloud types from visual",
        "Difficulty": "Medium",
        "Tags": "Weather, Meteorology",
    },
    {
        "Q. NO": "Q008",
        "Question Type": "Textual",
        "Transcript": "",
        "Instructions": "Read the passage and answer the question",
        "Question": "According to the passage, what was the primary cause of World War I",
        "Options": "",
        "Correct Answer": "The assassination of Archduke Franz Ferdinand",
        "Explanation": "The assassination of Archduke Franz Ferdinand of Austria triggered a chain of events including the mobilisation of European alliances which escalated into the full-scale war.",
        "Schema": "Short answer",
        "Question Purpose": "Test reading comprehension",
        "Difficulty": "Hard",
        "Tags": "History, WWI",
    },
    {
        "Q. NO": "Q009",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "Choose the correct answer.",
        "Question": "What is the result of 15 multiplied by 7?",
        "Options": "1) 95 | 2) 105 | 3) 115 | 4) 125",
        "Correct Answer": "2) 105",
        "Explanation": "When you multiply 15 by 7 you get 115 because 15 times 7 equals 115.",
        "Schema": "Single correct",
        "Question Purpose": "Test basic multiplication",
        "Difficulty": "Easy",
        "Tags": "Math, Arithmetic",
    },
    {
        "Q. NO": "Q010",
        "Question Type": "Image Based with Audio",
        "Transcript": "The narrator describes a busy marketplace scene with vendors selling fresh produce, colorful fabrics, and handmade crafts.",
        "Instructions": "Listen to the audio description and look at the image, then answer the question.",
        "Question": "Which item is NOT mentioned in the audio description?",
        "Options": "",
        "Correct Answer": "Electronic devices",
        "Explanation": "The audio only mentions produce, fabrics, and crafts. Electronic devices are not part of the described scene.",
        "Schema": "Short answer",
        "Question Purpose": "Test integrated audio-visual comprehension",
        "Difficulty": "Medium",
        "Tags": "Listening, Comprehension",
    },
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Questions"

header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for col_idx, col_name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_data in ROWS:
    ws.append([row_data.get(col, "") for col in COLUMNS])

col_widths = [8, 22, 35, 25, 50, 40, 20, 50, 15, 30, 10, 20]
from openpyxl.utils import get_column_letter
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
wb.save("test_input.xlsx")
print("Created test_input.xlsx with 10 test questions.")
print("Deliberate errors:")
print("  Q002: grammar error 'Dolphins is' + 'which of the following are'")
print("  Q003: ambiguous blank (missing units — degrees Celsius or Fahrenheit?)")
print("  Q005: question asks about efficiency but transcript doesn't say which is most efficient")
print("  Q006: informal language ('kinda') in question stem")
print("  Q007: missing question mark at end of question")
print("  Q008: missing question mark + missing period in instructions")
print("  Q009: options use numbers (1/2/3/4) instead of letters (A/B/C/D) + explanation contradicts answer (says 115 not 105)")
