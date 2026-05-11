import sqlite3
import uuid
import os
from datetime import datetime, timezone

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "feedback.db")


def _conn():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def _ensure_table():
    with _conn() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS feedback (
                id TEXT PRIMARY KEY,
                job_id TEXT NOT NULL,
                question_no TEXT NOT NULL,
                rubric_name TEXT NOT NULL,
                ai_score TEXT NOT NULL,
                ai_correction TEXT,
                user_verdict TEXT NOT NULL,
                user_correction TEXT,
                user_comment TEXT,
                original_text TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        con.execute("CREATE INDEX IF NOT EXISTS idx_job ON feedback(job_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_verdict ON feedback(user_verdict)")
        con.commit()


def _ensure_verdicts_table():
    with _conn() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS question_verdicts (
                id TEXT PRIMARY KEY,
                job_id TEXT NOT NULL,
                question_no TEXT NOT NULL,
                overall_status TEXT,
                verdict TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(job_id, question_no)
            )
        """)
        con.execute("CREATE INDEX IF NOT EXISTS idx_qv_job ON question_verdicts(job_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_qv_verdict ON question_verdicts(verdict)")
        con.commit()


_ensure_table()
_ensure_verdicts_table()


def insert_feedback(data: dict) -> dict:
    record = {k: v for k, v in data.items()}
    record["id"] = str(uuid.uuid4())
    record["created_at"] = datetime.now(timezone.utc).isoformat()
    record.pop("id", None)  # let us set it
    row_id = str(uuid.uuid4())

    with _conn() as con:
        con.execute("""
            INSERT INTO feedback
                (id, job_id, question_no, rubric_name, ai_score, ai_correction,
                 user_verdict, user_correction, user_comment, original_text, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (
            row_id,
            record.get("job_id", ""),
            record.get("question_no", ""),
            record.get("rubric_name", ""),
            record.get("ai_score", ""),
            record.get("ai_correction"),
            record.get("user_verdict", ""),
            record.get("user_correction"),
            record.get("user_comment"),
            record.get("original_text"),
            record.get("created_at", ""),
        ))
        con.commit()
    return {"id": row_id, **record}


def get_training_examples(rubric_name: str = None) -> list:
    with _conn() as con:
        if rubric_name:
            rows = con.execute(
                "SELECT * FROM feedback WHERE user_verdict IN ('reject','override') AND rubric_name=?",
                (rubric_name,)
            ).fetchall()
        else:
            rows = con.execute(
                "SELECT * FROM feedback WHERE user_verdict IN ('reject','override')"
            ).fetchall()
    return [dict(r) for r in rows]


def get_feedback_by_job(job_id: str) -> list:
    with _conn() as con:
        rows = con.execute(
            "SELECT * FROM feedback WHERE job_id=? ORDER BY created_at DESC",
            (job_id,)
        ).fetchall()
    return [dict(r) for r in rows]


def get_feedback_stats() -> dict:
    with _conn() as con:
        rows = con.execute("SELECT user_verdict FROM feedback").fetchall()
    stats = {"accept": 0, "reject": 0, "override": 0, "total": len(rows)}
    for row in rows:
        v = row["user_verdict"]
        if v in stats:
            stats[v] += 1
    return stats


def save_question_verdict(job_id: str, question_no: str, overall_status: str, verdict: str) -> dict:
    row_id = str(uuid.uuid4())
    created_at = datetime.now(timezone.utc).isoformat()
    with _conn() as con:
        con.execute("""
            INSERT INTO question_verdicts (id, job_id, question_no, overall_status, verdict, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(job_id, question_no) DO UPDATE SET
                verdict=excluded.verdict,
                overall_status=excluded.overall_status,
                created_at=excluded.created_at
        """, (row_id, job_id, question_no, overall_status, verdict, created_at))
        con.commit()
    return {"id": row_id, "verdict": verdict}


def get_verdict_stats() -> dict:
    with _conn() as con:
        rows = con.execute("SELECT verdict, overall_status FROM question_verdicts").fetchall()
    stats = {"approve": 0, "flag": 0, "ignore": 0, "total": len(rows)}
    for row in rows:
        v = row["verdict"]
        if v in stats:
            stats[v] += 1
    return stats


def get_verdicts_by_job(job_id: str) -> dict:
    with _conn() as con:
        rows = con.execute(
            "SELECT question_no, verdict FROM question_verdicts WHERE job_id=?",
            (job_id,)
        ).fetchall()
    return {row["question_no"]: row["verdict"] for row in rows}


RUBRIC_COLS = [
    "R1_Grammatical_Accuracy", "R2_Spelling", "R3_Ambiguity",
    "R4_Functionality_Alignment", "R5_Instruction_Clarity", "R6_Academic_Language",
    "R7_Option_Explanation_Consistency", "R8_Readability",
    "R9_Formatting_Spacing", "R10_Punctuation", "R11_EN_Consistency",
]


def get_accuracy_report(upload_dir: str) -> dict:
    """
    Cross-reference reviewed xlsx files with feedback.db to compute AI accuracy.

    Logic per question:
      - If user gave ANY 'reject' feedback on a rubric → AI was wrong on that question
      - If user gave ONLY 'accept'/'override' feedback → AI was correct
      - No feedback → unverified

    Returns breakdown for:
      Approved by AI  → correct (confirmed) vs incorrect (user found missed issues)
      Flagged by AI   → correctly flagged vs incorrectly flagged (false positive)
    """
    import glob
    import openpyxl

    # Step 1: Read all (job_id, q_no) → overall_status from reviewed xlsx files
    question_status = {}  # (job_id, q_no) → overall_status
    pattern = os.path.join(upload_dir, "reviewed_*.xlsx")
    for path in glob.glob(pattern):
        try:
            job_id = os.path.basename(path).replace("reviewed_", "").replace(".xlsx", "")
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Overall_Status" not in headers or "Q. NO" not in headers:
                wb.close()
                continue
            status_idx = headers.index("Overall_Status")
            qno_idx = headers.index("Q. NO")
            for row in ws.iter_rows(min_row=2, values_only=True):
                q_no = str(row[qno_idx]).strip() if qno_idx < len(row) and row[qno_idx] else ""
                status = str(row[status_idx]).strip() if status_idx < len(row) and row[status_idx] else ""
                if q_no and status and status != "None":
                    question_status[(job_id, q_no)] = status
            wb.close()
        except Exception:
            continue

    # Step 2: Read all feedback from SQLite → per (job_id, q_no) collect verdicts
    question_feedback = {}  # (job_id, q_no) → set of verdicts
    with _conn() as con:
        rows = con.execute("SELECT job_id, question_no, user_verdict FROM feedback").fetchall()
    for row in rows:
        key = (row["job_id"], row["question_no"])
        question_feedback.setdefault(key, set()).add(row["user_verdict"])

    # Step 3: Cross-reference
    approved_correct   = 0  # AI approved + user confirmed
    approved_incorrect = 0  # AI approved + user rejected (AI missed real issue)
    flagged_correct    = 0  # AI flagged + user confirmed (AI was right)
    flagged_incorrect  = 0  # AI flagged + user rejected (AI false positive)
    approved_unverified = 0
    flagged_unverified  = 0

    for (job_id, q_no), status in question_status.items():
        verdicts = question_feedback.get((job_id, q_no), set())
        is_approved = status == "Approved"
        is_flagged  = status in ("Needs Review", "Rejected")

        if not verdicts:
            if is_approved:
                approved_unverified += 1
            elif is_flagged:
                flagged_unverified += 1
            continue

        has_reject = "reject" in verdicts

        if is_approved:
            if has_reject:
                approved_incorrect += 1   # user said AI missed a real issue
            else:
                approved_correct += 1
        elif is_flagged:
            if has_reject:
                flagged_incorrect += 1    # user said AI wrongly flagged
            else:
                flagged_correct += 1

    total_approved = approved_correct + approved_incorrect + approved_unverified
    total_flagged  = flagged_correct + flagged_incorrect + flagged_unverified

    return {
        "approved_by_agent": {
            "total": total_approved,
            "correct":     approved_correct,
            "incorrect":   approved_incorrect,
            "unverified":  approved_unverified,
        },
        "flagged_by_agent": {
            "total": total_flagged,
            "correctly_flagged":   flagged_correct,
            "incorrectly_flagged": flagged_incorrect,
            "unverified":          flagged_unverified,
        },
    }


def get_upload_stats(upload_dir: str) -> dict:
    """Scan all reviewed_*.xlsx files and return a detailed breakdown per question."""
    import glob
    import openpyxl
    stats = {
        "total": 0, "files": 0,
        # Overall_Status breakdown
        "approved": 0,
        "approved_no_changes": 0,   # Approved + all rubrics Pass/N/A
        "approved_with_changes": 0, # Approved + some rubrics Minor
        "needs_review": 0,
        "rejected": 0,
        "review_failed": 0,
    }
    pattern = os.path.join(upload_dir, "reviewed_*.xlsx")
    for path in glob.glob(pattern):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Overall_Status" not in headers:
                wb.close()
                continue
            status_idx = headers.index("Overall_Status")
            rubric_indices = [headers.index(c) for c in RUBRIC_COLS if c in headers]
            stats["files"] += 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if status_idx >= len(row):
                    continue
                val = str(row[status_idx]).strip() if row[status_idx] else ""
                if not val or val == "None":
                    continue
                stats["total"] += 1
                if val == "Approved":
                    stats["approved"] += 1
                    rubric_scores = [str(row[i]).strip() if i < len(row) and row[i] else "" for i in rubric_indices]
                    has_minor = any(s == "Minor" for s in rubric_scores)
                    if has_minor:
                        stats["approved_with_changes"] += 1
                    else:
                        stats["approved_no_changes"] += 1
                elif val == "Needs Review":
                    stats["needs_review"] += 1
                elif val == "Rejected":
                    stats["rejected"] += 1
                elif val == "Review Failed":
                    stats["review_failed"] += 1
            wb.close()
        except Exception:
            continue
    return stats
