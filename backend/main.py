import os
import json
import time
import uuid
import threading
from typing import Optional

from fastapi import FastAPI, BackgroundTasks, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from config import UPLOAD_DIR, init_supabase, configure_dspy
from excel.reader import read_and_validate
from excel.writer import write_output
from excel.template import generate_template
from pipeline.reviewer import review_dataframe
from feedback.store import (
    insert_feedback, get_feedback_by_job, get_feedback_stats,
    save_question_verdict, get_verdict_stats, get_verdicts_by_job,
    get_upload_stats, get_accuracy_report,
)
from feedback.optimizer import optimize_all_rubrics
from sheets.logger import log_job_results, log_feedback, get_performance_stats_from_sheets, get_job_results_from_sheets

app = FastAPI(title="Exam Content Reviewer", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000",
                   os.environ.get("FRONTEND_URL", "")],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

job_store: dict = {}
job_store_lock = threading.Lock()


@app.on_event("startup")
def on_startup():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    try:
        init_supabase()
    except Exception as e:
        print(f"Warning: Supabase init failed: {e}. Feedback features will be unavailable.")


def _update_job(job_id: str, **kwargs):
    with job_store_lock:
        if job_id in job_store:
            job_store[job_id].update(kwargs)


def _run_review_job(job_id: str, df, input_path: str):
    def on_progress(current: int, total: int, q_no: str):
        progress = int((current / total) * 100) if total > 0 else 0
        _update_job(
            job_id,
            current=current,
            total=total,
            current_question=str(q_no),
            progress=progress,
        )

    try:
        results = review_dataframe(df, on_progress)

        out_filename = f"reviewed_{job_id}.xlsx"
        out_path = os.path.join(UPLOAD_DIR, out_filename)
        write_output(df, results, out_path)

        serializable_results = []
        for r in results:
            clean = {k: v for k, v in r.items() if k != "_rubric_details"}
            serializable_results.append(clean)

        _update_job(
            job_id,
            status="done",
            results=serializable_results,
            output_path=out_path,
            progress=100,
        )
        log_job_results(job_id, serializable_results)

    except Exception as e:
        _update_job(job_id, status="failed", error=str(e))


@app.post("/api/upload")
async def upload_file(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="Only Excel files (.xlsx, .xls) are accepted")

    job_id = str(uuid.uuid4())
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    input_path = os.path.join(UPLOAD_DIR, f"input_{job_id}.xlsx")

    contents = await file.read()
    with open(input_path, "wb") as f:
        f.write(contents)

    df, errors = read_and_validate(input_path)
    if errors:
        os.remove(input_path)
        raise HTTPException(422, detail={"errors": errors})

    total = len(df)
    with job_store_lock:
        job_store[job_id] = {
            "status": "processing",
            "progress": 0,
            "total": total,
            "current": 0,
            "current_question": "",
            "results": None,
            "output_path": None,
            "error": None,
            "started_at": time.time(),
            "input_path": input_path,
        }

    background_tasks.add_task(_run_review_job, job_id, df, input_path)
    return {"job_id": job_id, "total_questions": total}


@app.get("/api/status/{job_id}")
def get_status(job_id: str):
    job = job_store.get(job_id)
    if not job:
        raise HTTPException(404, detail="Job not found")
    return {
        "job_id": job_id,
        "status": job["status"],
        "progress": job["progress"],
        "current": job["current"],
        "total": job["total"],
        "current_question": job["current_question"],
        "elapsed_seconds": round(time.time() - job["started_at"], 1),
        "error": job.get("error"),
    }


@app.get("/api/results/{job_id}")
def get_results(job_id: str):
    job = job_store.get(job_id)
    if not job:
        raise HTTPException(404, detail="Job not found")
    if job["status"] != "done":
        raise HTTPException(409, detail=f"Job not complete. Status: {job['status']}")
    return {"job_id": job_id, "results": job["results"]}


@app.get("/api/download/{job_id}")
def download_results(job_id: str):
    job = job_store.get(job_id)
    if not job:
        raise HTTPException(404, detail="Job not found")
    if job["status"] != "done" or not job.get("output_path"):
        raise HTTPException(409, detail="Results not ready yet")
    return FileResponse(
        job["output_path"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"reviewed_{job_id}.xlsx",
    )


class FeedbackPayload(BaseModel):
    job_id: str
    question_no: str
    rubric_name: str
    ai_score: str
    ai_correction: Optional[str] = None
    user_verdict: str
    user_correction: Optional[str] = None
    user_comment: Optional[str] = None
    original_text: Optional[str] = None


@app.post("/api/feedback")
def submit_feedback(payload: FeedbackPayload):
    if payload.user_verdict not in ("accept", "reject", "override"):
        raise HTTPException(400, detail="user_verdict must be accept, reject, or override")
    try:
        data = payload.model_dump()
        record = insert_feedback(data)
        log_feedback(data)
        return {"status": "saved", "id": record.get("id")}
    except Exception as e:
        print(f"[FEEDBACK ERROR] {type(e).__name__}: {e}")
        raise HTTPException(500, detail=f"Failed to save feedback: {e}")


@app.post("/api/optimize")
def trigger_optimization():
    try:
        outcome = optimize_all_rubrics()
        return {
            "status": "complete",
            "results": outcome.get("results", outcome) if isinstance(outcome, dict) else outcome,
            "prompt_diff": outcome.get("prompt_diff") if isinstance(outcome, dict) else None,
        }
    except Exception as e:
        raise HTTPException(500, detail=f"Optimization failed: {e}")


@app.get("/api/optimize/preview")
def preview_optimized_prompt():
    """Return the current few-shot examples baked into the optimized prompts."""
    from pipeline.modules import OPTIMIZED_PATH
    if not os.path.exists(OPTIMIZED_PATH):
        return {"status": "not_optimized", "batches": []}
    try:
        with open(OPTIMIZED_PATH, "r") as f:
            data = json.load(f)
        batches = []
        batch_labels = {
            "mq_text":      "Text Mechanics (Grammar, Spelling, Punctuation, EN)",
            "mq_content":   "Content Quality (Alignment, Instructions, Academic, Readability)",
            "mq_structure": "Structure (Options/Explanation, Formatting)",
            "mq_ambiguity": "Ambiguity",
        }
        for key, label in batch_labels.items():
            predictor = data.get("predict", {}).get(key) or data.get(key, {})
            demos = predictor.get("demos", [])
            examples = []
            for d in demos:
                try:
                    q = json.loads(d.get("questions_json", "[]"))
                    r = json.loads(d.get("results_json", "[]"))
                    examples.append({
                        "question": q[0].get("question", "")[:200] if q else "",
                        "scores": {k: v for k, v in (r[0].items() if r else {}) if k.endswith("_score")},
                    })
                except Exception:
                    continue
            batches.append({"batch": key, "label": label, "example_count": len(demos), "examples": examples})
        return {"status": "optimized", "batches": batches}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/template")
def download_template():
    import io
    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    generate_template(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exam_questions_template.xlsx"},
    )


@app.get("/api/feedback/stats")
def feedback_stats():
    try:
        return get_feedback_stats()
    except Exception as e:
        raise HTTPException(500, detail=str(e))


class VerdictPayload(BaseModel):
    job_id: str
    question_no: str
    overall_status: Optional[str] = None
    verdict: str


@app.post("/api/question-verdict")
def submit_question_verdict(payload: VerdictPayload):
    if payload.verdict not in ("approve", "flag", "ignore"):
        raise HTTPException(400, detail="verdict must be approve, flag, or ignore")
    try:
        record = save_question_verdict(
            payload.job_id, payload.question_no,
            payload.overall_status or "", payload.verdict
        )
        return {"status": "saved", **record}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/verdicts/{job_id}")
def get_job_verdicts(job_id: str):
    return get_verdicts_by_job(job_id)


@app.get("/api/stats")
def overall_stats():
    try:
        upload_stats = get_upload_stats(UPLOAD_DIR)
        verdict_stats = get_verdict_stats()
        accuracy = get_accuracy_report(UPLOAD_DIR)
        return {
            "review": upload_stats,
            "verdicts": verdict_stats,
            "accuracy": accuracy,
        }
    except Exception as e:
        raise HTTPException(500, detail=str(e))


def _compute_performance(upload_stats: dict, accuracy: dict, source: str) -> dict:
    """Shared logic: turn upload_stats + accuracy dicts into the full performance payload."""
    total = upload_stats.get("total", 0)
    total_files           = upload_stats.get("files", 0)
    approved_no_changes   = upload_stats.get("approved_no_changes", 0)
    approved_with_changes = upload_stats.get("approved_with_changes", 0)
    needs_review  = upload_stats.get("needs_review", 0)
    rejected      = upload_stats.get("rejected", 0)
    review_failed = upload_stats.get("review_failed", 0)

    acc_flagged = accuracy.get("flagged_by_agent", {})

    # Confirmed false positives from human feedback (AI flagged but reviewer disagreed)
    fp_confirmed = acc_flagged.get("incorrectly_flagged", 0)

    # ── Confusion matrix ─────────────────────────────────────────────────────
    # Base proxy (always from upload stats — never inflated by unverified count):
    #   TP = all flagged questions minus confirmed false positives
    #   FP = human-confirmed false alarms
    #   FN = approved questions that had Minor corrections (AI missed real issues)
    #   TN = approved questions with no corrections at all (AI was right to approve)
    #
    # This stays consistent whether 0 or 10,000 feedback records exist.
    # "Approved with minor fixes" is the honest proxy for issues the AI missed —
    # it doesn't depend on whether a human explicitly rated them.
    flagged_total = needs_review + rejected
    tp = flagged_total - fp_confirmed
    fp = fp_confirmed
    fn = approved_with_changes
    tn = approved_no_changes

    precision    = round(tp / (tp + fp) * 100, 1) if (tp + fp) > 0 else 0.0
    recall       = round(tp / (tp + fn) * 100, 1) if (tp + fn) > 0 else 0.0
    f1           = round(2 * precision * recall / (precision + recall) / 100, 2) if (precision + recall) > 0 else 0.0
    accuracy_pct = round((tp + tn) / (tp + tn + fp + fn) * 100, 1) if total > 0 else 0.0

    insights = []
    if total == 0:
        insights.append("No review data yet — upload and review questions to see metrics here.")
    else:
        if fp == 0 and tp > 0:
            insights.append("Perfect precision — every flagged question has been a real issue so far.")
        elif precision < 70:
            insights.append(f"Low precision ({precision}%) — many flagged questions appear to be false alarms.")

        if recall < 50:
            insights.append(
                f"Low recall ({recall}%) — {fn} questions with fixable issues were approved without flagging."
            )
        elif recall >= 80:
            insights.append(f"Strong recall ({recall}%) — most problematic questions are being caught.")

        if approved_with_changes > 0:
            pct = round(approved_with_changes / total * 100, 1)
            insights.append(
                f"{approved_with_changes} questions ({pct}%) were approved but had minor corrections applied — "
                "the flagging threshold may be too strict for minor issues."
            )

        if accuracy_pct >= 90:
            insights.append(f"High overall accuracy ({accuracy_pct}%) — the agent is broadly reliable.")
        elif accuracy_pct < 70:
            insights.append(f"Overall accuracy ({accuracy_pct}%) needs improvement — consider running the DSPy optimizer.")

    approved_total = approved_no_changes + approved_with_changes
    flagged_total  = needs_review + rejected

    return {
        "total_questions": total,
        "total_files": total_files,
        "data_source": source,
        "metrics": {"precision": precision, "recall": recall, "f1_score": f1, "accuracy": accuracy_pct},
        "confusion_matrix": {"tp": tp, "fp": fp, "fn": fn, "tn": tn},
        "question_outcomes": {
            "correct_approvals": tn,
            "missed_issues": fn,
            "correctly_flagged": tp,
            "false_flags": fp,
            "approved_no_changes": approved_no_changes,
            "approved_with_changes": approved_with_changes,
            "needs_review": needs_review,
            "rejected": rejected,
            "review_failed": review_failed,
        },
        "agent_summary": {
            "approved_total": approved_total,
            "approved_correct": approved_no_changes,
            "approved_with_fixes": approved_with_changes,
            "flagged_total": flagged_total,
            "flagged_needs_revision": needs_review,
            "flagged_rejected": rejected,
            "flagged_false_positive": fp_confirmed,
            "review_failed": review_failed,
        },
        "diagnostic_insights": insights,
    }


@app.get("/api/performance")
def performance_dashboard():
    """
    Compute precision/recall/F1/accuracy from Google Sheets (permanent store).
    Falls back to local reviewed_*.xlsx files if Sheets is unavailable.
    Results are cached for 5 minutes server-side.
    """
    try:
        sheet_data = get_performance_stats_from_sheets()
        if sheet_data:
            return _compute_performance(
                sheet_data["upload_stats"],
                sheet_data["accuracy"],
                source="Google Sheets",
            )
        # Fallback to local files
        upload_stats = get_upload_stats(UPLOAD_DIR)
        accuracy     = get_accuracy_report(UPLOAD_DIR)
        return _compute_performance(upload_stats, accuracy, source="local files")
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/sheets/results/{job_id}")
def get_results_from_sheets(job_id: str):
    """
    Fallback endpoint: fetch per-job results from Google Sheets.
    Used when the in-memory job store has expired (e.g. after a Render restart).
    """
    results = get_job_results_from_sheets(job_id)
    if results is None:
        raise HTTPException(404, detail="Job results not found in Google Sheets.")
    return {"job_id": job_id, "results": results, "source": "sheets"}


@app.get("/api/download-from-sheets/{job_id}")
def download_from_sheets(job_id: str):
    """
    Regenerate a reviewed Excel file from Google Sheets data and return it for download.
    Used as a fallback when the original upload file has expired from the ephemeral disk.
    """
    results = get_job_results_from_sheets(job_id)
    if results is None:
        raise HTTPException(404, detail="Job results not found in Google Sheets.")

    import tempfile
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    RUBRIC_COLS = [
        "R1_Grammatical_Accuracy", "R2_Spelling", "R3_Ambiguity",
        "R4_Functionality_Alignment", "R5_Instruction_Clarity", "R6_Academic_Language",
        "R7_Option_Explanation_Consistency", "R8_Readability",
        "R9_Formatting_Spacing", "R10_Punctuation", "R11_EN_Consistency",
    ]
    SCORE_FILLS = {
        "Pass":     PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Minor":    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Major":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }
    STATUS_FILLS = {
        "Approved":      PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Needs Review":  PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Rejected":      PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Review Failed": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    }

    # Column order for the regenerated sheet
    DISPLAY_COLS = [
        "Q. NO", "Question Type", "Question", "Correct Answer", "Difficulty",
        "Overall_Status",
        *RUBRIC_COLS,
        "Remarks",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Review Results (from Sheets)"

    headers = [c for c in DISPLAY_COLS if any(c in r for r in results)]
    if not headers:
        headers = DISPLAY_COLS

    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 28

    for row_idx, r in enumerate(results, 2):
        for col_idx, h in enumerate(headers, 1):
            val = r.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if h in RUBRIC_COLS and val in SCORE_FILLS:
                cell.fill = SCORE_FILLS[val]
            elif h == "Overall_Status" and val in STATUS_FILLS:
                cell.fill = STATUS_FILLS[val]

    for col_idx, h in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(r.get(h, "") or "")) for r in results),
            default=0
        )
        ws.column_dimensions[col_letter].width = min(max(len(h), max_len) + 2, 60)

    ws.freeze_panes = "A2"

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    out_path = os.path.join(UPLOAD_DIR, f"sheets_export_{job_id[:8]}.xlsx")
    wb.save(out_path)

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"reviewed_{job_id[:8]}.xlsx",
    )


class ExportQuestionsPayload(BaseModel):
    questions: list[dict]


@app.post("/api/export-questions")
def export_questions(payload: ExportQuestionsPayload):
    import io
    import re
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from config import REQUIRED_COLUMNS

    _label_re = re.compile(r"^[A-Za-z]\)\s*")

    def normalize_options(val):
        s = str(val).strip() if val else ""
        if not s:
            return ""
        if " | " in s and _label_re.search(s):
            return "\n".join(_label_re.sub("", p.strip()) for p in s.split(" | "))
        return s

    COL_WIDTHS = {
        "Q. NO": 8, "Question Type": 22, "Transcript": 45, "Instructions": 40,
        "Question": 55, "Options": 30, "Correct Answer": 25,
        "Explanation": 55, "Schema": 18, "Question Purpose": 35,
        "Difficulty": 12, "Tags": 30,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Updated Questions"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, header in enumerate(REQUIRED_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_num, q in enumerate(payload.questions, 2):
        for col_idx, col in enumerate(REQUIRED_COLUMNS, 1):
            val = q.get(col, "") or ""
            if col == "Options":
                val = normalize_options(val)
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 50

    for col_idx, header in enumerate(REQUIRED_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 20)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=updated_questions.xlsx"},
    )


@app.get("/health")
def health():
    return {"status": "ok"}
