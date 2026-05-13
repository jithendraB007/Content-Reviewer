import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import REQUIRED_COLUMNS

# ── Sample rows — one per use case ───────────────────────────────────────────

SAMPLE_ROWS = [
    {
        "_note": "Use Case 1: Standard MCQ — Instruction + Question + Options + Answer",
        "Q. NO": "Q001",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "Choose the correct answer.",
        "Question": "What is the capital of France?",
        "Options": "Paris\nLondon\nBerlin\nMadrid",
        "Correct Answer": "Paris",
        "Explanation": "Paris is the capital and largest city of France.",
        "Schema": "Geography",
        "Question Purpose": "Test knowledge of European capitals",
        "Difficulty": "Easy",
        "Tags": "Geography, Europe, Capitals",
    },
    {
        "_note": "Use Case 2: No Instruction — Question stands alone (leave Instructions blank)",
        "Q. NO": "Q002",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "",
        "Question": "Which planet is closest to the Sun?",
        "Options": "Mercury\nVenus\nEarth\nMars",
        "Correct Answer": "Mercury",
        "Explanation": "Mercury is the first planet from the Sun in our solar system.",
        "Schema": "Science",
        "Question Purpose": "Test basic astronomy knowledge",
        "Difficulty": "Easy",
        "Tags": "Science, Solar System",
    },
    {
        "_note": "Use Case 3: No Question — Instruction only (leave Question blank)",
        "Q. NO": "Q003",
        "Question Type": "Fill in the Blanks",
        "Transcript": "",
        "Instructions": "Fill in the blank with the correct verb form: She ______ to school every day.",
        "Question": "",
        "Options": "",
        "Correct Answer": "goes",
        "Explanation": "The subject 'She' is singular, so the verb takes 'goes' in simple present tense.",
        "Schema": "Grammar",
        "Question Purpose": "Test subject-verb agreement",
        "Difficulty": "Easy",
        "Tags": "Grammar, Verb Forms",
    },
    {
        "_note": "Use Case 4: Audio / Listening — Transcript + Instruction + Question + Options",
        "Q. NO": "Q004",
        "Question Type": "Audio Based MCQ",
        "Transcript": "The speaker discussed the importance of time management in the workplace. She said that employees who plan their day in advance are more productive.",
        "Instructions": "Listen to the audio and choose the correct answer.",
        "Question": "What does the speaker say about employees who plan their day?",
        "Options": "They are more productive\nThey earn higher salaries\nThey work fewer hours\nThey avoid meetings",
        "Correct Answer": "They are more productive",
        "Explanation": "The transcript explicitly states that employees who plan their day in advance are more productive.",
        "Schema": "Listening Comprehension",
        "Question Purpose": "Test listening and inference skills",
        "Difficulty": "Medium",
        "Tags": "Listening, Workplace, Time Management",
    },
    {
        "_note": "Use Case 5: Fill in the Blanks — Instruction + Question with blank, no Options needed",
        "Q. NO": "Q005",
        "Question Type": "Fill in the Blanks",
        "Transcript": "",
        "Instructions": "Fill in the blank with the most appropriate word.",
        "Question": "The team worked ______ to finish the project before the deadline.",
        "Options": "",
        "Correct Answer": "hard",
        "Explanation": "'Hard' means diligently. 'Hardly' means rarely — a common confusion.",
        "Schema": "Vocabulary",
        "Question Purpose": "Test word usage and adverb vs adjective distinction",
        "Difficulty": "Medium",
        "Tags": "Grammar, Vocabulary, Adverbs",
    },
    {
        "_note": "Use Case 6: Merged Question+Options — entire question block in the Question column",
        "Q. NO": "Q006",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": "Choose the correct answer.",
        "Question": "Which of the following is a renewable energy source?\n1. Coal\n2. Natural Gas\n3. Solar Power\n4. Petroleum",
        "Options": "",
        "Correct Answer": "Solar Power",
        "Explanation": "Solar power is renewable because it comes from the Sun, which is an inexhaustible source of energy.",
        "Schema": "Environment",
        "Question Purpose": "Test knowledge of energy sources",
        "Difficulty": "Easy",
        "Tags": "Environment, Energy, Science",
    },
]

# Row background colours — alternating light tints per use case
ROW_FILLS = [
    PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid"),  # light blue
    PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),  # light green
    PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),  # light yellow
    PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid"),  # light purple
    PatternFill(start_color="FDFEFE", end_color="FDFEFE", fill_type="solid"),  # near white
    PatternFill(start_color="FEF5E7", end_color="FEF5E7", fill_type="solid"),  # light orange
]

COL_WIDTHS = {
    "Q. NO": 8, "Question Type": 22, "Transcript": 45, "Instructions": 40,
    "Question": 55, "Options": 30, "Correct Answer": 25,
    "Explanation": 55, "Schema": 18, "Question Purpose": 35,
    "Difficulty": 12, "Tags": 30,
}

USE_CASE_GUIDE = [
    ("USE CASE", "INSTRUCTIONS", "QUESTION", "OPTIONS", "CORRECT ANSWER", "WHEN TO USE"),
    (
        "Standard MCQ",
        "Fill in",
        "Fill in",
        "One option per line\n(no A/B/C/D labels)",
        "Exact text of correct option\n(no A/B/C/D prefix)",
        "Regular multiple-choice questions",
    ),
    (
        "No Instruction",
        "Leave BLANK",
        "Fill in",
        "One option per line",
        "Exact text",
        "Question is self-explanatory — no separate instruction needed",
    ),
    (
        "No Question",
        "Full task goes here\n(e.g. 'Fill in the blank: She ___')",
        "Leave BLANK",
        "Leave blank if FIB",
        "Correct word / phrase",
        "Fill in the Blanks or tasks where instruction IS the question",
    ),
    (
        "Audio / Listening",
        "Fill in",
        "Fill in",
        "One option per line",
        "Exact text",
        "Audio Based MCQ — put the audio script in Transcript",
    ),
    (
        "Fill in the Blanks",
        "Fill in",
        "Include ____ for the blank",
        "Leave BLANK",
        "The missing word / phrase",
        "No options needed — just the blank marker in the question",
    ),
    (
        "Merged Question+Options",
        "Fill in or leave blank",
        "Put everything here\n(question text + all choices merged)",
        "Leave BLANK",
        "Exact text of correct option",
        "When your source has question and choices in one block",
    ),
]


def generate_template(output_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, header in enumerate(REQUIRED_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Sample rows ───────────────────────────────────────────────────────────
    for row_num, (sample, fill) in enumerate(zip(SAMPLE_ROWS, ROW_FILLS), 2):
        for col_idx, col in enumerate(REQUIRED_COLUMNS, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=sample.get(col, ""))
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_num].height = 60

    # Add a "Note" column after the required columns to label each use case
    note_col = len(REQUIRED_COLUMNS) + 1
    note_header = ws.cell(row=1, column=note_col, value="Note (delete before uploading)")
    note_header.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    note_header.font = Font(bold=True, color="888888", size=10, italic=True)
    note_header.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(note_col)].width = 38

    for row_num, (sample, fill) in enumerate(zip(SAMPLE_ROWS, ROW_FILLS), 2):
        note_cell = ws.cell(row=row_num, column=note_col, value=sample.get("_note", ""))
        note_cell.font = Font(color="666666", italic=True, size=9)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, header in enumerate(REQUIRED_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 20)

    ws.freeze_panes = "A2"

    # ── Instructions sheet ────────────────────────────────────────────────────
    info_ws = wb.create_sheet(title="How to Fill")

    title_font   = Font(bold=True, size=14, color="1E3A5F")
    heading_font = Font(bold=True, size=11)
    body_font    = Font(size=10)
    note_font    = Font(size=9, italic=True, color="666666")
    hdr_fill     = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    hdr_font     = Font(bold=True, color="FFFFFF", size=10)

    info_ws["A1"] = "EXAM CONTENT REVIEWER — How to Fill the Template"
    info_ws["A1"].font = title_font
    info_ws.row_dimensions[1].height = 24

    info_ws["A3"] = "GENERAL RULES"
    info_ws["A3"].font = heading_font

    rules = [
        "• Options: list each option on a separate line — NO A) B) C) D) labels, NO pipe ( | ) separators.",
        "• Correct Answer: write the exact text of the correct option — NO A) prefix.",
        "• Instructions OR Question must be filled — both can be filled, but not both empty.",
        "• Transcript: fill only for Audio Based MCQ / Image Based with Audio. Leave blank otherwise.",
        "• Options: leave blank for Fill in the Blanks and when options are merged inside the Question.",
        "• Delete the 'Note' column (last column) before uploading the file.",
    ]
    for i, rule in enumerate(rules, 4):
        cell = info_ws.cell(row=i, column=1, value=rule)
        cell.font = body_font
        info_ws.row_dimensions[i].height = 18

    # Use-case table
    info_ws.cell(row=11, column=1, value="USE CASE REFERENCE TABLE").font = heading_font
    info_ws.row_dimensions[11].height = 20

    col_letters = ["A", "B", "C", "D", "E", "F"]
    col_widths_info = [22, 22, 22, 28, 28, 40]

    for col_idx, (letter, width) in enumerate(zip(col_letters, col_widths_info), 1):
        info_ws.column_dimensions[letter].width = width

    for row_offset, row_data in enumerate(USE_CASE_GUIDE):
        excel_row = 12 + row_offset
        for col_idx, value in enumerate(row_data, 1):
            cell = info_ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_offset == 0:
                cell.fill = hdr_fill
                cell.font = hdr_font
            else:
                cell.font = body_font
                fill_color = ROW_FILLS[(row_offset - 1) % len(ROW_FILLS)]
                cell.fill = fill_color
        info_ws.row_dimensions[excel_row].height = 45

    # Question Types section
    qt_row = 12 + len(USE_CASE_GUIDE) + 2
    info_ws.cell(row=qt_row, column=1, value="ACCEPTED QUESTION TYPES").font = heading_font
    types = [
        "MCQ", "Fill in the Blanks", "Speaking Based", "Textual",
        "Audio Based MCQ", "Prompt Based",
        "Image Based with Options", "Image Based with Audio",
    ]
    for i, t in enumerate(types, qt_row + 1):
        info_ws.cell(row=i, column=1, value=f"  • {t}").font = body_font

    diff_row = qt_row + len(types) + 2
    info_ws.cell(row=diff_row, column=1, value="DIFFICULTY VALUES").font = heading_font
    info_ws.cell(row=diff_row + 1, column=1, value="  Easy   |   Medium   |   Hard").font = body_font

    if isinstance(output_path, str) and os.path.dirname(output_path):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
