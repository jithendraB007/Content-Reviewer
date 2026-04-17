import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

ADDED_COLUMNS = [
    "Corrected_Instructions",
    "Corrected_Question",
    "Corrected_Options",
    "Corrected_Explanation",
    "Corrected_Transcript",
    "R1_Grammatical_Accuracy",
    "R2_Spelling",
    "R3_Ambiguity",
    "R4_Functionality_Alignment",
    "R5_Instruction_Clarity",
    "R6_Academic_Language",
    "R7_Option_Explanation_Consistency",
    "R8_Readability",
    "R9_Formatting_Spacing",
    "R10_Punctuation",
    "R11_EN_Consistency",
    "Overall_Status",
    "Remarks",
]

RUBRIC_COLUMNS = [c for c in ADDED_COLUMNS if c.startswith("R") and "_" in c]
STATUS_COLUMN = "Overall_Status"

SCORE_FILLS = {
    "Pass": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Minor": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Major": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Critical": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}

STATUS_FILLS = {
    "Approved": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "Needs Review": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "Rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "Review Failed": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
}

BOLD_HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")


def write_output(original_df: pd.DataFrame, results: list, output_path: str) -> None:
    results_df = pd.DataFrame(results)

    for col in ADDED_COLUMNS:
        if col not in results_df.columns:
            results_df[col] = ""

    combined = pd.concat(
        [original_df.reset_index(drop=True), results_df[ADDED_COLUMNS].reset_index(drop=True)],
        axis=1,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Review Results"

    headers = list(combined.columns)
    orig_col_count = len(original_df.columns)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_idx > orig_col_count:
            cell.font = BOLD_HEADER_FONT
            cell.fill = HEADER_FILL
        else:
            cell.font = Font(bold=True)
    ws.row_dimensions[1].height = 28

    for row_idx, row_data in enumerate(combined.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    num_data_rows = len(combined)
    last_data_row = num_data_rows + 1

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_data_row}"

        if header in RUBRIC_COLUMNS:
            for score, fill in SCORE_FILLS.items():
                rule = FormulaRule(
                    formula=[f'EXACT({col_letter}2,"{score}")'],
                    fill=fill,
                    stopIfTrue=False,
                )
                ws.conditional_formatting.add(cell_range, rule)

        elif header == STATUS_COLUMN:
            for status, fill in STATUS_FILLS.items():
                rule = FormulaRule(
                    formula=[f'EXACT({col_letter}2,"{status}")'],
                    fill=fill,
                    stopIfTrue=False,
                )
                ws.conditional_formatting.add(cell_range, rule)

    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_idx in range(2, last_data_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(cell_val)), 60))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 62)

    ws.freeze_panes = "A2"

    _write_summary_sheet(wb, results)

    os.makedirs(os.path.dirname(output_path), exist_ok=True) if os.path.dirname(output_path) else None
    wb.save(output_path)


def _write_summary_sheet(wb: Workbook, results: list) -> None:
    ws = wb.create_sheet(title="Summary")

    total = len(results)
    approved = sum(1 for r in results if r.get("Overall_Status") == "Approved")
    needs_review = sum(1 for r in results if r.get("Overall_Status") == "Needs Review")
    rejected = sum(1 for r in results if r.get("Overall_Status") == "Rejected")
    failed = sum(1 for r in results if r.get("Overall_Status") == "Review Failed")

    def pct(n):
        return f"{n / total * 100:.1f}%" if total else "0%"

    title_font = Font(bold=True, size=13)
    header_font = Font(bold=True)
    ws["A1"] = "Review Summary"
    ws["A1"].font = title_font
    ws.append([])

    ws.append(["Status", "Count", "Percentage"])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for label, count in [
        ("Total Questions", total),
        ("Approved", approved),
        ("Needs Review", needs_review),
        ("Rejected", rejected),
        ("Review Failed", failed),
    ]:
        ws.append([label, count, pct(count) if label != "Total Questions" else "100%"])

    ws.append([])
    ws.append(["Rubric", "Pass", "Minor", "Major", "Critical", "N/A", "Skipped"])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for rubric in RUBRIC_COLUMNS:
        counts = {s: 0 for s in ("Pass", "Minor", "Major", "Critical", "N/A")}
        skipped = 0
        for r in results:
            score = r.get(rubric, "")
            if score in counts:
                counts[score] += 1
            elif score == "" or score is None:
                skipped += 1
        ws.append([
            rubric,
            counts["Pass"], counts["Minor"], counts["Major"],
            counts["Critical"], counts["N/A"], skipped,
        ])

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 35 if col == "A" else 10
