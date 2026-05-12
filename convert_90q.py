"""
Converts 90-question CSV into content reviewer Excel template format.
Usage: python convert_90q.py [input.csv] [output.xlsx]
"""

import re
import sys
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

CSV_PATH = sys.argv[1] if len(sys.argv) > 1 else "input_90q.csv"
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else "90q_reviewer_input.xlsx"

HTML_TAG_RE  = re.compile(r"<[^>]+>")
# Split sentences that run together: "Opt A.Opt B" or "Opt A. Opt B"
SENT_SPLIT_RE = re.compile(r"(?<=[.!?])\s*(?=[A-Z])")


def fix_encoding(text):
    """Reverse double mojibake (cp1252 + utf-8 pass x2)."""
    if not isinstance(text, str):
        return ""
    try:
        return text.encode("cp1252").decode("utf-8").encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    try:
        return text.encode("cp1252").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    try:
        return text.encode("latin-1").decode("utf-8")
    except (UnicodeDecodeError, UnicodeEncodeError):
        pass
    return text


def strip_html(text):
    text = HTML_TAG_RE.sub(" ", text)
    return re.sub(r" {2,}", " ", text).strip()


def get_question_type(instruction):
    if "fill in the blank" in instruction.lower():
        return "Fill in the Blanks"
    return "MCQ"


def split_options(raw):
    """
    Return a list of up to 5 option strings from the raw field.
    Handles: newline-separated, pipe-separated, or concatenated sentences.
    """
    if not raw:
        return []
    raw = raw.strip()

    # Newline-separated (most rows)
    if "\n" in raw:
        return [o.strip() for o in raw.split("\n") if o.strip()]

    # Pipe-separated
    if "|" in raw:
        return [o.strip() for o in raw.split("|") if o.strip()]

    # Sentences glued together: split on sentence boundary before a capital letter
    parts = SENT_SPLIT_RE.split(raw)
    if len(parts) >= 2:
        return [p.strip() for p in parts if p.strip()]

    # Fallback: single option or indeterminate
    return [raw]


def format_options(raw):
    """Convert raw options string -> 'A) opt1 | B) opt2 | C) opt3 | D) opt4'."""
    opts = split_options(str(raw))[:5]
    if not opts:
        return ""
    labels = ["A", "B", "C", "D", "E"]
    return " | ".join(f"{labels[i]}) {opt}" for i, opt in enumerate(opts))


def extract_module(exclusive_tag):
    t = str(exclusive_tag).upper()
    if t.endswith("_GRAMMAR"):
        return "Grammar"
    if t.endswith("_VOCABULARY"):
        return "Vocabulary"
    if "VERBAL_ABILITY" in t:
        return "Verbal Ability"
    last = exclusive_tag.split("_")[-1]
    return last.replace("-", " ").title() if last else ""


# Read CSV
df = pd.read_csv(CSV_PATH, encoding="utf-8-sig", keep_default_na=False)
print(f"Loaded {len(df)} rows from {CSV_PATH}")

rows = []
for _, row in df.iterrows():
    q_no        = str(row.get("Q. NO", "")).strip()
    instruction = fix_encoding(str(row.get("Instruction", "")).strip())
    question_raw = fix_encoding(str(row.get("Question", "")).strip())
    options_raw  = fix_encoding(str(row.get("Options", "")).strip())
    answer       = fix_encoding(str(row.get("Answer", "")).strip())
    explanation  = fix_encoding(str(row.get("Explanation", "")).strip())
    complexity   = str(row.get("Complexity", "")).strip()
    excl_tag     = str(row.get("Exclusive Tag", "")).strip()

    question_clean = strip_html(question_raw)
    q_type         = get_question_type(instruction)
    options_fmt    = format_options(options_raw)
    schema         = extract_module(excl_tag)

    rows.append({
        "Q. NO":          q_no,
        "Question Type":  q_type,
        "Transcript":     "",
        "Instructions":   instruction,
        "Question":       question_clean,
        "Options":        options_fmt,
        "Correct Answer": answer,
        "Explanation":    explanation,
        "Schema":         schema,
        "Question Purpose": instruction,
        "Difficulty":     complexity,
        "Tags":           excl_tag,
    })

out_df = pd.DataFrame(rows)

# Write Excel
with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    out_df.to_excel(writer, index=False, sheet_name="Questions")
    ws = writer.sheets["Questions"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    col_widths = {
        "A": 8,  "B": 20, "C": 12, "D": 50,
        "E": 60, "F": 45, "G": 35, "H": 55,
        "I": 15, "J": 45, "K": 10, "L": 50,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"\nWritten {len(out_df)} rows -> {OUT_PATH}")
print("\nQuestion Type distribution:")
print(out_df["Question Type"].value_counts().to_string())
print("\nDifficulty distribution:")
print(out_df["Difficulty"].value_counts().to_string())
print("\nSchema distribution:")
print(out_df["Schema"].value_counts().to_string())

# Warn about option counts that are not 4
import warnings
print("\nOption count check:")
bad = 0
for i, r in out_df.iterrows():
    opts = str(r["Options"]).split(" | ")
    if len(opts) not in (4, 5) and str(r["Options"]).strip():
        print(f"  {r['Q. NO']}: {len(opts)} options - {str(r['Options'])[:60]}")
        bad += 1
if bad == 0:
    print("  All rows have 4 options OK")
