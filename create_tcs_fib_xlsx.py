"""
Converts TCS NQT Verbal Ability - Fill in the Blanks CSV into reviewer input XLSX.
Run: python create_tcs_fib_xlsx.py
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

RAW_DATA = [
    # (Instruction, Question, Options list, Correct Answer, Explanation, Difficulty)
    ("Fill in the blank with the most appropriate option.",
     "She ______ to the office every day.",
     ["go", "goes", "going", "gone"],
     "goes", "Singular subject 'she' takes 'goes'.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "They are looking ______ the new project.",
     ["at", "for", "to", "on"],
     "at", "'Looking at' means examining something.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "He is good ______ solving problems.",
     ["at", "in", "on", "for"],
     "at", "Correct preposition is 'good at'.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "We will start the meeting ______ 10 AM.",
     ["in", "on", "at", "by"],
     "at", "'At' is used for specific time.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "She has ______ completed her assignment.",
     ["yet", "already", "still", "just"],
     "already", "'Already' shows completion before now.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "She was ______ tired to continue working.",
     ["so", "too", "very", "much"],
     "too", "'Too...to' structure indicates excess preventing action.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "The manager spoke ______ to ensure everyone understood the instructions.",
     ["clear", "clearly", "clearest", "clearing"],
     "clearly", "Adverb needed to modify the verb 'spoke'.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "He has been working here ______ five years.",
     ["since", "for", "from", "by"],
     "for", "'For' is used with duration.", "Easy"),

    ("Fill in the blank with the most appropriate option.",
     "The team worked ______ to complete the project before the deadline.",
     ["hardly", "hard", "harder", "hardest"],
     "hard", "'Hard' means diligently; 'hardly' means rarely.", "Medium"),

    ("Fill in the blank with the most appropriate option.",
     "If she ______ earlier, she would have caught the train.",
     ["leaves", "left", "had left", "was leaving"],
     "had left", "Third conditional requires past perfect.", "Medium"),
]


def fmt_options(opts):
    labels = ["A", "B", "C", "D"]
    return " | ".join(f"{labels[i]}) {o}" for i, o in enumerate(opts))


rows = []
for i, (instr, question, opts, answer, expl, diff) in enumerate(RAW_DATA, 1):
    rows.append({
        "Q. NO": f"Q{i:03d}",
        "Question Type": "Fill in the Blanks",
        "Transcript": "",
        "Instructions": instr,
        "Question": question,
        "Options": fmt_options(opts),
        "Correct Answer": answer,
        "Explanation": expl,
        "Schema": "",
        "Question Purpose": "Fill in the Blanks",
        "Difficulty": diff,
        "Tags": "TCS NQT, Verbal Ability, Fill in the Blanks",
    })

df = pd.DataFrame(rows)
out_path = "tcs_nqt_fib_input.xlsx"

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Questions")
    ws = writer.sheets["Questions"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    col_widths = {
        "A": 8, "B": 20, "C": 12, "D": 40, "E": 60,
        "F": 40, "G": 30, "H": 50, "I": 12, "J": 25, "K": 10, "L": 35,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Created {out_path} with {len(df)} questions")
print(df["Difficulty"].value_counts().to_string())
