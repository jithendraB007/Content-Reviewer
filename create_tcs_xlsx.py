"""
Converts TCS NQT Verbal Ability - Reading Comprehension CSV into reviewer input XLSX.
Run: python create_tcs_xlsx.py
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

RAW_DATA = [
    # (Instruction, Question, Options list, Correct Answer, Explanation, Difficulty)
    ("Read the passage and answer the following question.",
     "Digital learning has grown rapidly in recent years. It allows students to access educational content anytime and anywhere. However, it also requires self-discipline and motivation, which many learners struggle with. Institutions are now combining online and offline methods to improve learning outcomes.\n\nWhat is the main idea of the passage?",
     ["Digital learning is ineffective.", "Digital learning has benefits and challenges.", "Students dislike online education.", "Offline learning is outdated."],
     "Digital learning has benefits and challenges.", "The passage highlights both advantages and challenges.", "Easy"),

    ("Read the passage and answer the following question.",
     "Digital learning has grown rapidly in recent years. It allows students to access educational content anytime and anywhere. However, it also requires self-discipline and motivation, which many learners struggle with. Institutions are now combining online and offline methods to improve learning outcomes.\n\nWhat is one challenge of digital learning?",
     ["Lack of teachers", "High cost", "Need for self-discipline", "Limited access to content"],
     "Need for self-discipline", "Clearly mentioned in the passage.", "Easy"),

    ("Read the passage and answer the following question.",
     "Digital learning has grown rapidly in recent years. It allows students to access educational content anytime and anywhere. However, it also requires self-discipline and motivation, which many learners struggle with. Institutions are now combining online and offline methods to improve learning outcomes.\n\nWhy are institutions combining methods?",
     ["To reduce costs", "To improve learning outcomes", "To eliminate digital learning", "To increase workload"],
     "To improve learning outcomes", "Direct inference from last sentence.", "Medium"),

    ("Read the passage and answer the following question.",
     "Workplace communication plays a crucial role in organizational success. Clear and concise communication ensures that tasks are completed efficiently. Miscommunication, on the other hand, can lead to delays and confusion. Companies are investing in training programs to enhance employees' communication skills.\n\nWhat does the passage emphasize?",
     ["Importance of communication", "Role of technology", "Employee salaries", "Office design"],
     "Importance of communication", "Central idea of passage.", "Easy"),

    ("Read the passage and answer the following question.",
     "Workplace communication plays a crucial role in organizational success. Clear and concise communication ensures that tasks are completed efficiently. Miscommunication, on the other hand, can lead to delays and confusion. Companies are investing in training programs to enhance employees' communication skills.\n\nWhat is a result of miscommunication?",
     ["Increased efficiency", "Delays and confusion", "Better teamwork", "Higher profits"],
     "Delays and confusion", "Explicitly stated.", "Easy"),

    ("Read the passage and answer the following question.",
     "Workplace communication plays a crucial role in organizational success. Clear and concise communication ensures that tasks are completed efficiently. Miscommunication, on the other hand, can lead to delays and confusion. Companies are investing in training programs to enhance employees' communication skills.\n\nWhat does the word 'concise' most nearly mean?",
     ["Lengthy.", "Clear and brief.", "Confusing.", "Detailed."],
     "Clear and brief.", "'Concise' refers to expressing information clearly in few words.", "Medium"),

    ("Read the passage and answer the following question.",
     "Artificial intelligence (AI) is increasingly being used in workplaces to improve efficiency and decision-making. By automating repetitive tasks, AI allows employees to focus on more strategic work. However, concerns about job displacement and ethical use of data remain significant. Organizations are therefore trying to balance innovation with responsible implementation.\n\nWhat is the main purpose of using AI in workplaces?",
     ["To replace all employees.", "To improve efficiency and decision-making.", "To reduce salaries.", "To eliminate technology."],
     "To improve efficiency and decision-making.", "The passage clearly states AI is used to improve efficiency and decision-making.", "Easy"),

    ("Read the passage and answer the following question.",
     "Artificial intelligence (AI) is increasingly being used in workplaces to improve efficiency and decision-making. By automating repetitive tasks, AI allows employees to focus on more strategic work. However, concerns about job displacement and ethical use of data remain significant. Organizations are therefore trying to balance innovation with responsible implementation.\n\nWhat is one benefit of AI mentioned in the passage?",
     ["Increased workload.", "Automation of repetitive tasks.", "Reduced innovation.", "Less productivity."],
     "Automation of repetitive tasks.", "Directly mentioned as a key benefit.", "Easy"),

    ("Read the passage and answer the following question.",
     "Artificial intelligence (AI) is increasingly being used in workplaces to improve efficiency and decision-making. By automating repetitive tasks, AI allows employees to focus on more strategic work. However, concerns about job displacement and ethical use of data remain significant. Organizations are therefore trying to balance innovation with responsible implementation.\n\nWhat concern is associated with AI?",
     ["Lack of usage.", "Job displacement and ethical issues.", "High employee satisfaction.", "Improved communication."],
     "Job displacement and ethical issues.", "Mentioned as a major concern in the passage.", "Medium"),

    ("Read the passage and answer the following question.",
     "Time management is a critical skill in today's fast-paced work environment. Employees who prioritize tasks effectively are more likely to meet deadlines and achieve goals. Poor time management can lead to stress, missed opportunities, and decreased productivity. To address this, many organizations encourage employees to use planning tools and set clear priorities.\n\nWhat is the central idea of the passage?",
     ["Importance of time management.", "Use of technology.", "Workplace design.", "Employee salaries."],
     "Importance of time management.", "The passage focuses on why time management matters.", "Easy"),

    ("Read the passage and answer the following question.",
     "Time management is a critical skill in today's fast-paced work environment. Employees who prioritize tasks effectively are more likely to meet deadlines and achieve goals. Poor time management can lead to stress, missed opportunities, and decreased productivity. To address this, many organizations encourage employees to use planning tools and set clear priorities.\n\nWhat happens due to poor time management?",
     ["Increased productivity.", "Better performance.", "Stress and missed opportunities.", "Improved planning."],
     "Stress and missed opportunities.", "Clearly stated consequence.", "Easy"),

    ("Read the passage and answer the following question.",
     "Time management is a critical skill in today's fast-paced work environment. Employees who prioritize tasks effectively are more likely to meet deadlines and achieve goals. Poor time management can lead to stress, missed opportunities, and decreased productivity. To address this, many organizations encourage employees to use planning tools and set clear priorities.\n\nWhy do organizations encourage planning tools?",
     ["To increase workload.", "To help employees manage time effectively.", "To reduce communication.", "To eliminate deadlines."],
     "To help employees manage time effectively.", "Inference based on final sentence.", "Medium"),

    ("Read the passage and answer the following question.",
     "Teamwork is essential for achieving organizational goals. When individuals collaborate effectively, they can combine their skills and ideas to produce better outcomes. However, lack of coordination and communication can lead to misunderstandings and reduced efficiency. Companies often conduct team-building activities to strengthen collaboration.\n\nWhat is the main idea of the passage?",
     ["Teamwork is unnecessary.", "Teamwork improves outcomes but requires coordination.", "Employees prefer working alone.", "Communication is not important."],
     "Teamwork improves outcomes but requires coordination.", "Passage highlights both benefits and challenges.", "Easy"),

    ("Read the passage and answer the following question.",
     "Teamwork is essential for achieving organizational goals. When individuals collaborate effectively, they can combine their skills and ideas to produce better outcomes. However, lack of coordination and communication can lead to misunderstandings and reduced efficiency. Companies often conduct team-building activities to strengthen collaboration.\n\nWhat is one benefit of teamwork?",
     ["Reduced efficiency.", "Better outcomes through collaboration.", "Increased misunderstandings.", "Lack of coordination."],
     "Better outcomes through collaboration.", "Direct detail from passage.", "Easy"),

    ("Read the passage and answer the following question.",
     "Teamwork is essential for achieving organizational goals. When individuals collaborate effectively, they can combine their skills and ideas to produce better outcomes. However, lack of coordination and communication can lead to misunderstandings and reduced efficiency. Companies often conduct team-building activities to strengthen collaboration.\n\nWhat does the word 'collaborate' most nearly mean?",
     ["Compete.", "Work together.", "Ignore others.", "Lead alone."],
     "Work together.", "Context suggests working jointly.", "Medium"),

    ("Read the passage and answer the following question.",
     "Customer satisfaction plays a key role in the success of any business. When customers feel valued and receive quality service, they are more likely to remain loyal. On the other hand, poor service can lead to negative feedback and loss of business. Companies therefore focus on improving customer experience through training and feedback systems.\n\nWhat is the central idea of the passage?",
     ["Customer satisfaction is important for business success.", "Customers do not matter.", "Training is unnecessary.", "Feedback systems are ineffective."],
     "Customer satisfaction is important for business success.", "Main focus of passage.", "Easy"),

    ("Read the passage and answer the following question.",
     "Customer satisfaction plays a key role in the success of any business. When customers feel valued and receive quality service, they are more likely to remain loyal. On the other hand, poor service can lead to negative feedback and loss of business. Companies therefore focus on improving customer experience through training and feedback systems.\n\nWhat happens when customer service is poor?",
     ["Increased loyalty.", "Better feedback.", "Loss of business.", "Higher profits."],
     "Loss of business.", "Clearly mentioned.", "Easy"),

    ("Read the passage and answer the following question.",
     "Customer satisfaction plays a key role in the success of any business. When customers feel valued and receive quality service, they are more likely to remain loyal. On the other hand, poor service can lead to negative feedback and loss of business. Companies therefore focus on improving customer experience through training and feedback systems.\n\nWhat does the word 'loyal' most nearly mean?",
     ["Disloyal.", "Faithful.", "Angry.", "Careless."],
     "Faithful.", "Loyal = continuing support/trust.", "Medium"),

    ("Read the passage and answer the following question.",
     "Leadership is the ability to guide and influence others toward achieving common goals. Effective leaders communicate clearly, make informed decisions, and motivate their teams. Poor leadership, however, can result in confusion and lack of direction. Organizations invest in leadership development programs to build strong leaders.\n\nWhat is the main idea of the passage?",
     ["Leadership is unnecessary.", "Leadership is important for guiding teams effectively.", "Teams do not need direction.", "Decision-making is irrelevant."],
     "Leadership is important for guiding teams effectively.", "Focus is on leadership importance.", "Easy"),

    ("Read the passage and answer the following question.",
     "Leadership is the ability to guide and influence others toward achieving common goals. Effective leaders communicate clearly, make informed decisions, and motivate their teams. Poor leadership, however, can result in confusion and lack of direction. Organizations invest in leadership development programs to build strong leaders.\n\nWhat is one quality of effective leaders?",
     ["Poor communication.", "Clear communication.", "Lack of motivation.", "Indecision."],
     "Clear communication.", "Directly stated.", "Easy"),

    ("Read the passage and answer the following question.",
     "Leadership is the ability to guide and influence others toward achieving common goals. Effective leaders communicate clearly, make informed decisions, and motivate their teams. Poor leadership, however, can result in confusion and lack of direction. Organizations invest in leadership development programs to build strong leaders.\n\nWhat does the word 'influence' most nearly mean?",
     ["Ignore.", "Control or affect.", "Follow.", "Avoid."],
     "Control or affect.", "Influence = ability to affect others.", "Medium"),
]


def fmt_options(opts):
    labels = ["A", "B", "C", "D"]
    return " | ".join(f"{labels[i]}) {o}" for i, o in enumerate(opts))


rows = []
for i, (instr, question, opts, answer, expl, diff) in enumerate(RAW_DATA, 1):
    # Split passage from question text
    parts = question.split("\n\n", 1)
    q_text = parts[1].strip() if len(parts) == 2 else question
    passage = parts[0].strip() if len(parts) == 2 else ""

    rows.append({
        "Q. NO": f"Q{i:03d}",
        "Question Type": "MCQ",
        "Transcript": "",
        "Instructions": instr,
        "Question": f"{passage}\n\n{q_text}" if passage else q_text,
        "Options": fmt_options(opts),
        "Correct Answer": answer,
        "Explanation": expl,
        "Schema": "",
        "Question Purpose": "Reading Comprehension",
        "Difficulty": diff,
        "Tags": "TCS NQT, Verbal Ability, Reading Comprehension",
    })

df = pd.DataFrame(rows)
out_path = "tcs_nqt_input.xlsx"

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Questions")
    ws = writer.sheets["Questions"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"

    col_widths = {
        "A": 8, "B": 20, "C": 12, "D": 40, "E": 60,
        "F": 40, "G": 30, "H": 50, "I": 12, "J": 25, "K": 10, "L": 35,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

print(f"Created {out_path} with {len(df)} questions")
print(df["Difficulty"].value_counts().to_string())
